# -*- coding: utf-8 -*-
"""
批量解析raw_file_dir目录下的doc文件，参照keyword.txt中配置的关键字，解析其值，将解析结果输出为excel文件
"""
import os
import re
import docx
import openpyxl


# 获取docx 文件内容
def read_line_from_docx(file_path):
    document = docx.Document(file_path)
    all_paragraphs = document.paragraphs
    paragraphs_text = []
    for paragraph in all_paragraphs:
        paragraphs_text.append(paragraph.text)
    return paragraphs_text


# 在文件内容中解析所有key对应的值
def find_key_in_str(keyword, pg_text):
    match = re.search(rf"{keyword}(\w+)[\u4e00-\u9fa5,.]", pg_text)
    if match:
        return match.group(1)
    else:
        return None


# 创建一组字典，将查询出的结果写入字典中
def create_last_arr_list(keyword_list, file_path):
    global all_value
    my_dict = {}
    parag_text = read_line_from_docx(file_path)
    for keyWord in keyword_list:
        for pag_text in parag_text:
            all_value = find_key_in_str(keyWord, pag_text)
            if all_value is not None:
                my_dict[keyWord] = all_value
                break
            else:
                my_dict[keyWord] = "未解析到内容"
    return my_dict


# 将汇总完成的字典写入表格
def load_data_in_excel(keyword__list, file__path, output__filename):
    file__list = get_files_in_directory(file__path)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    line_num = 2

    # 将字典的键作为表头写入第一行
    for file_path in file__list:
        print("开始解析%s文件" % file_path)
        data = create_last_arr_list(keyword__list, file_path)
        header_row = 1
        for col_num, header_title in enumerate(data.keys(), start=1):
            worksheet.cell(row=header_row, column=col_num).value = header_title
        # 将字典的值写入表格
        data_row = line_num
        for col_num, cell_value in enumerate(data.values(), start=1):
            worksheet.cell(row=data_row, column=col_num).value = cell_value
        line_num += 1

    # 写入完成，保存表格
    workbook.save(output__filename)


# 解析目录下所有文件列表，并返回文件路径
def get_files_in_directory(directory):
    raw_file_list = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            raw_file_list.append(os.path.join(root, file))
    return raw_file_list


# 获取待查找关键字文件，解析其内容
def read_txt(txt_path):
    file = open(txt_path, "r", encoding="utf-8")
    file_contents = file.readline()
    data = file_contents
    key_list = data.split(";")
    return key_list


if __name__ == "__main__":
    # 录入需要查找的字段，用;隔开，需要在模板文件中唯一
    read_keyword_list = read_txt("keyword.txt")
    # 需要解析的文件目录
    path = "raw_file_dir"
    output_filename = "测试文档输出.xlsx"
    file_list = get_files_in_directory(path)
    print(file_list)
    load_data_in_excel(read_keyword_list, path, output_filename)
